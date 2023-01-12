#!/usr/bin/env python3
 
import argparse
import sys
import pandas as pd
from openpyxl.utils import get_column_letter

def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("-i", "--input-file", help="Filepath of csv or tsv file to read from", required=True)
    parser.add_argument("-o", "--output-file", help="Filepath of xlsx file to write create", required=True)
    return parser.parse_args()

def resize_columns(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value

if __name__ == '__main__':
    args = parse_args()
    separator = ','
    if args.input_file.endswith('tsv'):
        separator = '\t'
    cvs_data_frame = pd.read_csv(args.input_file, sep=separator)
    excel_file = pd.ExcelWriter(args.output_file)
    cvs_data_frame.to_excel(excel_file, index=False, freeze_panes=(0,1))
    ws = excel_file.sheets['Sheet1']
    max_column_letter = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f'A1:{max_column_letter}{ws.max_row}'
    resize_columns(ws)
    excel_file.save()